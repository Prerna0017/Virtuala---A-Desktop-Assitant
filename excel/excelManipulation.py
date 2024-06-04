import openpyxl

def singleCellRead():
    path = "excel/trial.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj.cell(row = 1, column = 1)
    print(cell_obj.value)

def multiCellRead():
    path = "excel/trial.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    print("Total Rows:", row)
    print("Total Columns:", column)
    print("\nValue of first column")
    for i in range(1, row + 1): 
        cell_obj = sheet_obj.cell(row = i, column = 1) 
        print(cell_obj.value)
    print("\nValue of first row")
    for i in range(1, column + 1): 
        cell_obj = sheet_obj.cell(row = 2, column = i) 
        print(cell_obj.value, end = " ")

def multiSliceCellRead():
    path = "excel/trial.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    cell_obj = sheet_obj['A1': 'B6']
    for cell1, cell2 in cell_obj:
        print(cell1.value, cell2.value)

def writeCell():
    wb = openpyxl.Workbook()
    sheet = wb.active
    c1 = sheet.cell(row = 1, column = 1)
    c1.value = "Hello"
    c2 = sheet.cell(row= 1 , column = 2)
    c2.value = "World"
    c3 = sheet['A2']
    c3.value = "Welcome"
    c4 = sheet['B2']
    c4.value = "Everyone"
    wb.save("excel/sample.xlsx")

def appendCell():
    wb = openpyxl.load_workbook("excel/data.xlsx")
    sheet = wb.active
    data = [
        (1, 'gmail', 'prerna', 'pass123')
    ]
    for row in data:
        sheet.append(row)
    wb.save('excel/data.xlsx')



# singleCellRead()
# multiCellRead()
# multiSliceCellRead()
# writeCell()
appendCell()